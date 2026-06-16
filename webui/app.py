#!/usr/bin/env python3
"""
Model Benchmark WebUI - Flask Backend
Manages test execution, concurrency, results storage, and API endpoints.
"""

import json
import os
import sys
import subprocess
import io
import threading
import time
import uuid
import signal
from datetime import datetime
from pathlib import Path
import urllib.request
import urllib.error

# Fix Windows GBK encoding
os.environ.setdefault('PYTHONUTF8', '1')
try:
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
except (ValueError, AttributeError):
    pass

from flask import Flask, jsonify, request, send_file, Response
from flask_cors import CORS

# ── Paths ──────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEST_SPEC_PATH = PROJECT_ROOT / "test_spec.json"
CONFIG_PATH = PROJECT_ROOT / "config.json"
RESULTS_DIR = PROJECT_ROOT / "results"
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
STATIC_DIR = Path(__file__).resolve().parent / "static"

RESULTS_DIR.mkdir(parents=True, exist_ok=True)

# ── Flask App ──────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=str(STATIC_DIR))
CORS(app)

@app.before_request
def check_auth():
    if request.path == '/' or request.path.startswith('/static/') or request.path == '/favicon.ico':
        return None
    
    access_password = CONFIG.get("access_password")
    if not access_password:
        return None
        
    auth_header = request.headers.get("X-Access-Password")
    auth_query = request.args.get("password")
    
    if auth_header == access_password or auth_query == access_password:
        return None
        
    return jsonify({"error": "Unauthorized", "message": "需要访问密码"}), 401

@app.after_request
def gzip_response(response):
    # Set Cache-Control for static resources to enable long-term browser caching (1 year), excluding HTML
    if request.path.startswith('/static/') and not request.path.endswith('.html'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'

    accept_encoding = request.headers.get('Accept-Encoding', '')
    if 'gzip' not in accept_encoding.lower():
        return response
    if response.status_code < 200 or response.status_code >= 300:
        return response
    if response.headers.get('Content-Encoding'):
        return response
    
    content_type = response.headers.get('Content-Type', '')
    if 'json' in content_type or 'html' in content_type or 'javascript' in content_type or 'css' in content_type:
        response.direct_passthrough = False
        data = response.get_data()
        if len(data) > 500:
            import gzip
            gzip_buffer = io.BytesIO()
            with gzip.GzipFile(mode='wb', compresslevel=6, fileobj=gzip_buffer) as gzip_file:
                gzip_file.write(data)
            response.set_data(gzip_buffer.getvalue())
            response.headers['Content-Encoding'] = 'gzip'
            response.headers['Content-Length'] = len(response.get_data())
            response.headers['Vary'] = 'Accept-Encoding'
    return response

# ── State Management ───────────────────────────────────────────────────────
# Load test spec and config
with open(TEST_SPEC_PATH, "r", encoding="utf-8") as f:
    TEST_SPEC = json.load(f)

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    CONFIG = json.load(f)

# Build flat test registry
TESTS = {}
for suite in TEST_SPEC.get("test_suites", []):
    for test in suite.get("tests", []):
        test["_suite_id"] = suite["id"]
        test["_suite_name"] = suite["name"]
        test["_suite_parallel"] = suite.get("parallel", True)
        # Effective parallel flag: test-level overrides suite-level
        test["_effective_parallel"] = test.get("parallel", suite.get("parallel", True))
        TESTS[test["id"]] = test

# Test execution state
class TestRunner:
    """Manages test execution with concurrency control."""
    
    def __init__(self):
        self.lock = threading.Lock()
        self.non_parallel_lock = threading.Lock()  # Mutex for non-parallel tests
        self.running_tests = {}  # test_id -> {process, status, progress, output, ...}
        self.completed_tests = {}  # test_id -> result dict
        self._load_existing_results()
    
    def _load_existing_results(self):
        """Load previously saved results from disk. Skip stale/incomplete results."""
        for result_file in RESULTS_DIR.glob("*.json"):
            if "_progress" in result_file.stem:
                continue  # Skip progress files
            try:
                with open(result_file, "r", encoding="utf-8") as f:
                    result = json.load(f)
                test_id = result_file.stem
                # Skip stale results: must have test_id AND actual content (details/metrics/result)
                if "test_id" not in result:
                    continue
                # Check if result has actual content (not just {test_id, status, result: {}})
                has_content = (
                    result.get("details") or
                    result.get("metrics") or
                    (isinstance(result.get("result"), dict) and bool(result.get("result")))
                )
                if not has_content:
                    result_file.unlink()  # Delete stale result file
                    continue
                
                # Standardize raw CLI results by wrapping them into a consistent structure
                if "result" not in result:
                    test = TESTS.get(test_id, {})
                    result = {
                        "test_id": test_id,
                        "test_name": test.get("name", test_id),
                        "suite": test.get("_suite_name", ""),
                        "status": result.get("status", "passed"),
                        "result": {
                            "details": result.get("details", ""),
                            "metrics": result.get("metrics", {}),
                        },
                        "completed_at": result.get("end_time") or result.get("completed_at") or datetime.now().isoformat()
                    }
                    with open(result_file, "w", encoding="utf-8") as StandardFile:
                        json.dump(result, StandardFile, ensure_ascii=False, indent=2)

                self.completed_tests[test_id] = result
            except (json.JSONDecodeError, IOError):
                pass
    
    def _sync_progress_from_file(self, test_id):
        """Sync progress from the results/{test_id}_progress.json file."""
        progress_file = RESULTS_DIR / f"{test_id}_progress.json"
        if not progress_file.exists():
            return
        try:
            with open(progress_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            current_step = data.get("current_step")
            progress_pct = data.get("progress_pct")
            
            with self.lock:
                if test_id in self.running_tests:
                    rt = self.running_tests[test_id]
                    if current_step:
                        rt["message"] = current_step
                    if progress_pct is not None:
                        try:
                            p = int(float(progress_pct))
                            p = min(max(p, 0), 95)
                            rt["progress"] = max(rt.get("progress", 0), p)
                        except (ValueError, TypeError):
                            pass
        except Exception:
            pass

    def get_all_statuses(self):
        """Get status of all tests."""
        for test_id in TESTS:
            is_running = False
            with self.lock:
                is_running = test_id in self.running_tests
            if is_running:
                self._sync_progress_from_file(test_id)

        statuses = {}
        with self.lock:
            for test_id in TESTS:
                if test_id in self.running_tests:
                    statuses[test_id] = {
                        "status": "running",
                        "progress": self.running_tests[test_id].get("progress", 0),
                        "message": self.running_tests[test_id].get("message", ""),
                        "started_at": self.running_tests[test_id].get("started_at"),
                    }
                elif test_id in self.completed_tests:
                    statuses[test_id] = {
                        "status": self.completed_tests[test_id].get("status", "completed"),
                        "progress": 100,
                        "completed_at": self.completed_tests[test_id].get("completed_at"),
                        "result": self.completed_tests[test_id].get("result"),
                    }
                else:
                    statuses[test_id] = {"status": "pending", "progress": 0}
        return statuses
    
    def get_test_status(self, test_id):
        """Get status of a specific test."""
        is_running = False
        with self.lock:
            is_running = test_id in self.running_tests
        if is_running:
            self._sync_progress_from_file(test_id)

        with self.lock:
            if test_id in self.running_tests:
                rt = self.running_tests[test_id]
                return {
                    "status": "running",
                    "progress": rt.get("progress", 0),
                    "message": rt.get("message", ""),
                    "output_lines": rt.get("output_lines", [])[-50:],  # last 50 lines
                    "started_at": rt.get("started_at"),
                }
            elif test_id in self.completed_tests:
                return {
                    "status": self.completed_tests[test_id].get("status", "completed"),
                    "progress": 100,
                    "completed_at": self.completed_tests[test_id].get("completed_at"),
                    "result": self.completed_tests[test_id].get("result"),
                }
            else:
                return {"status": "pending", "progress": 0}
    
    def is_non_parallel_running(self):
        """Check if any non-parallel test is currently running."""
        with self.lock:
            for tid, rt in self.running_tests.items():
                if not TESTS[tid].get("_effective_parallel", True):
                    return True
        return False
    
    def can_run(self, test_id):
        """Check if a test can be started."""
        test = TESTS.get(test_id)
        if not test:
            return False, "Test not found"
        with self.lock:
            if test_id in self.running_tests:
                return False, "Test is already running"
        if not test.get("_effective_parallel", True):
            if self.is_non_parallel_running():
                return False, "Another non-parallel test is running (mutex lock)"
        return True, "OK"
    
    def start_test(self, test_id):
        """Start a test execution in a background thread. Clears old results first."""
        can_run, reason = self.can_run(test_id)
        if not can_run:
            return False, reason
        
        test = TESTS[test_id]
        
        # Clear old results before re-running
        with self.lock:
            if test_id in self.completed_tests:
                del self.completed_tests[test_id]
        result_file = RESULTS_DIR / f"{test_id}.json"
        progress_file = RESULTS_DIR / f"{test_id}_progress.json"
        for f in [result_file, progress_file]:
            if f.exists():
                f.unlink()
        
        with self.lock:
            self.running_tests[test_id] = {
                "progress": 0,
                "message": "Initializing...",
                "output_lines": [],
                "started_at": datetime.now().isoformat(),
                "process": None,
            }
        
        thread = threading.Thread(
            target=self._run_test_worker,
            args=(test_id, test),
            daemon=True,
        )
        thread.start()
        return True, "Test started"
    
    def _run_test_worker(self, test_id, test):
        """Worker function that executes a test."""
        script_path = SCRIPTS_DIR / f"{test_id}.py"
        runner_path = SCRIPTS_DIR / "runner.py"
        
        try:
            with self.lock:
                self.running_tests[test_id]["message"] = "Starting test..."
                self.running_tests[test_id]["progress"] = 5
            
            if script_path.exists():
                # Run individual test script directly
                self._execute_script(test_id, script_path)
            elif runner_path.exists():
                # Use unified runner.py with --test-id
                self._execute_runner(test_id, runner_path)
            else:
                # Simulate test execution for demo purposes
                self._simulate_test(test_id, test)
            
        except Exception as e:
            self._finish_test(test_id, "failed", {
                "error": str(e),
                "message": f"Test failed with exception: {e}",
            })
    
    def _execute_script(self, test_id, script_path):
        """Execute a Python test script and capture output."""
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env["TEST_ID"] = test_id
        env["PROJECT_ROOT"] = str(PROJECT_ROOT)
        
        # Add project root to PATH for claude CLI
        env["PATH"] = f"/Users/langqi/.npm-global/bin:{env.get('PATH', '')}"
        
        env["PYTHONUTF8"] = "1"
        try:
            proc = subprocess.Popen(
                [sys.executable, str(script_path)],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                encoding='utf-8', errors='replace',
                env=env,
                cwd=str(PROJECT_ROOT),
            )
            
            with self.lock:
                self.running_tests[test_id]["process"] = proc
            
            output_lines = []
            for line in iter(proc.stdout.readline, ""):
                line = line.strip()
                if line:
                    output_lines.append(line)
                    self._sync_progress_from_file(test_id)
                    with self.lock:
                        if test_id in self.running_tests:
                            self.running_tests[test_id]["output_lines"] = output_lines[-50:]
                            self.running_tests[test_id]["message"] = line
                            # Parse progress from output if available
                            if "PROGRESS:" in line:
                                try:
                                    p = int(line.split("PROGRESS:")[1].strip().rstrip("%"))
                                    self.running_tests[test_id]["progress"] = max(self.running_tests[test_id].get("progress", 0), min(p, 95))
                                except (ValueError, IndexError):
                                    pass
                            else:
                                # Also parse tqdm progress bars like Evaluating[...]:  X%| or Running[eval]:  X%|
                                import re
                                if "eval" in line.lower() or "running" in line.lower():
                                    matches = re.findall(r'\b([0-9]+)%\s*\|', line)
                                    if matches:
                                        try:
                                            p = int(matches[-1])
                                            if any(k in line.lower() for k in ["download", "pulling", "downloading", ".tar", ".zip", ".json", ".gz"]):
                                                p_mapped = 5 + int(p * 0.1)
                                            else:
                                                p_mapped = p
                                            self.running_tests[test_id]["progress"] = max(self.running_tests[test_id].get("progress", 0), min(p_mapped, 95))
                                        except ValueError:
                                            pass
                            if "MESSAGE:" in line:
                                try:
                                    self.running_tests[test_id]["message"] = line.split("MESSAGE:", 1)[1].strip()
                                except IndexError:
                                    pass
            
            proc.wait(timeout=3600)  # 1 hour max
            
            if proc.returncode == 0:
                # Try to load result from results directory
                result_file = RESULTS_DIR / f"{test_id}.json"
                if result_file.exists():
                    with open(result_file, "r", encoding="utf-8") as f:
                        result_data = json.load(f)
                    # runner.py outputs: {test_id, status, details, metrics, end_time}
                    status = result_data.get("status", "passed")
                    # Preserve full result: include details + metrics, not just "result"
                    full_result = {
                        "details": result_data.get("details", ""),
                        "metrics": result_data.get("metrics", {}),
                    }
                    self._finish_test(test_id, status, full_result)
                else:
                    self._finish_test(test_id, "passed", {
                        "details": "\n".join(output_lines[-20:]),
                        "returncode": 0,
                    })
            else:
                self._finish_test(test_id, "failed", {
                    "details": "\n".join(output_lines[-20:]),
                    "returncode": proc.returncode,
                    "message": f"Script exited with code {proc.returncode}",
                })
                
        except subprocess.TimeoutExpired:
            proc.kill()
            self._finish_test(test_id, "failed", {"message": "Test timed out (1 hour limit)"})
        except Exception as e:
            self._finish_test(test_id, "failed", {"message": str(e)})
    
    def _execute_runner(self, test_id, runner_path):
        """Execute test via unified runner.py --test-id."""
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env["TEST_ID"] = test_id
        env["PROJECT_ROOT"] = str(PROJECT_ROOT)
        env["PATH"] = f"/Users/langqi/.npm-global/bin:{env.get('PATH', '')}"
        
        # Load config for API credentials
        config_path = PROJECT_ROOT / "config.json"
        api_base = ""
        api_key = ""
        if config_path.exists():
            with open(config_path, "r") as f:
                cfg = json.load(f)
            api_base = cfg.get("model", {}).get("api_base", "")
            api_key = cfg.get("model", {}).get("api_key", "") or os.environ.get(cfg.get("model", {}).get("api_key_env", ""), "")
        
        cmd = [sys.executable, str(runner_path), "--test-id", test_id]
        if api_base:
            cmd.extend(["--api-base", api_base])
        if api_key:
            cmd.extend(["--api-key", api_key])
        
        env["PYTHONUTF8"] = "1"
        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                encoding='utf-8', errors='replace',
                env=env,
                cwd=str(PROJECT_ROOT),
            )
            
            with self.lock:
                self.running_tests[test_id]["process"] = proc
            
            output_lines = []
            for line in iter(proc.stdout.readline, ""):
                line = line.strip()
                if line:
                    output_lines.append(line)
                    with self.lock:
                        if test_id in self.running_tests:
                            self.running_tests[test_id]["output_lines"] = output_lines[-50:]
                            self.running_tests[test_id]["message"] = line
                            if "PROGRESS:" in line:
                                        try:
                                            p = int(line.split("PROGRESS:")[1].strip().rstrip("%"))
                                            self.running_tests[test_id]["progress"] = min(p, 95)
                                        except (ValueError, IndexError):
                                            pass
                            else:
                                # Also parse tqdm progress bars like Evaluating[...]:  X%| or Running[eval]:  X%|
                                import re
                                if "eval" in line.lower() or "running" in line.lower():
                                    matches = re.findall(r'\b([0-9]+)%\s*\|', line)
                                    if matches:
                                        try:
                                            p = int(matches[-1])
                                            self.running_tests[test_id]["progress"] = min(p, 95)
                                        except ValueError:
                                            pass
                            if "MESSAGE:" in line:
                                try:
                                    self.running_tests[test_id]["message"] = line.split("MESSAGE:", 1)[1].strip()
                                except IndexError:
                                    pass
            
            proc.wait(timeout=3600)
            
            result_file = RESULTS_DIR / f"{test_id}.json"
            if proc.returncode == 0:
                if result_file.exists():
                    with open(result_file, "r", encoding="utf-8") as f:
                        result_data = json.load(f)
                    status = result_data.get("status", "passed")
                    full_result = {
                        "details": result_data.get("details", ""),
                        "metrics": result_data.get("metrics", {}),
                    }
                    self._finish_test(test_id, status, full_result)
                else:
                    self._finish_test(test_id, "passed", {
                        "details": "\n".join(output_lines[-20:]),
                        "returncode": 0,
                    })
            else:
                self._finish_test(test_id, "failed", {
                    "details": "\n".join(output_lines[-20:]),
                    "returncode": proc.returncode,
                    "message": f"Script exited with code {proc.returncode}",
                })
        except subprocess.TimeoutExpired:
            proc.kill()
            self._finish_test(test_id, "failed", {"message": "Test timed out (1 hour limit)"})
        except Exception as e:
            self._finish_test(test_id, "failed", {"message": str(e)})
    
    def _simulate_test(self, test_id, test):
        """Simulate test execution for demo/testing purposes."""
        import random
        
        test_type = test.get("type", "api_call")
        total_steps = random.randint(10, 30)
        
        # Simulate duration based on test type
        step_delay = {
            "api_call": 0.5,
            "benchmark": 1.5,
            "perf": 2.0,
            "load_test": 2.5,
            "streaming": 1.0,
            "comprehensive": 1.5,
            "monitor": 0.3,
        }.get(test_type, 1.0)
        
        messages = [
            "Connecting to API endpoint...",
            "Sending test requests...",
            "Validating response format...",
            "Checking parameter defaults...",
            "Running assertion checks...",
            "Collecting metrics...",
            "Verifying results...",
            "Computing statistics...",
            "Generating report...",
        ]
        
        for step in range(total_steps):
            time.sleep(step_delay * random.uniform(0.5, 1.5))
            progress = int((step + 1) / total_steps * 90) + 5
            
            with self.lock:
                if test_id not in self.running_tests:
                    return  # Test was cancelled
                self.running_tests[test_id]["progress"] = progress
                self.running_tests[test_id]["message"] = messages[step % len(messages)]
                self.running_tests[test_id]["output_lines"].append(
                    f"[Step {step+1}/{total_steps}] {messages[step % len(messages)]}"
                )
        
        # Determine result
        has_official = test.get("official") is not None
        if has_official:
            official = test["official"]
            tolerance = test.get("tolerance", 4.0)
            score = official + random.uniform(-tolerance * 1.2, tolerance * 0.8)
            passed = abs(score - official) <= tolerance
            result = {
                "score": round(score, 2),
                "official": official,
                "tolerance": tolerance,
                "delta": round(score - official, 2),
                "passed": passed,
            }
            status = "passed" if passed else "failed"
        elif test_type == "perf":
            # Performance test simulation
            ttft_p50 = random.uniform(1.0, 5.0)
            ttft_p90 = random.uniform(3.0, 12.0)
            result = {
                "ttft_p50": round(ttft_p50, 2),
                "ttft_p90": round(ttft_p90, 2),
                "otps": round(random.uniform(15, 50), 1),
                "success_rate": round(random.uniform(98, 100), 2),
            }
            status = "passed" if random.random() > 0.2 else "failed"
        else:
            status = "passed" if random.random() > 0.15 else "failed"
            result = {
                "checks_passed": random.randint(3, 10),
                "checks_total": 10,
                "details": f"Test completed with status: {status}",
            }
        
        self._finish_test(test_id, status, result)
    
    def _finish_test(self, test_id, status, result):
        """Mark a test as completed and save results."""
        with self.lock:
            if test_id in self.running_tests:
                del self.running_tests[test_id]
        
        completed_at = datetime.now().isoformat()
        test = TESTS[test_id]
        
        result_data = {
            "test_id": test_id,
            "test_name": test["name"],
            "suite": test["_suite_name"],
            "status": status,
            "result": result,
            "completed_at": completed_at,
        }
        
        with self.lock:
            self.completed_tests[test_id] = result_data
        
        # Save to disk
        result_file = RESULTS_DIR / f"{test_id}.json"
        with open(result_file, "w", encoding="utf-8") as f:
            json.dump(result_data, f, ensure_ascii=False, indent=2)
    
    def cancel_test(self, test_id):
        """Cancel a running test and clean up files."""
        with self.lock:
            if test_id in self.running_tests:
                proc = self.running_tests[test_id].get("process")
                if proc:
                    proc.terminate()
                del self.running_tests[test_id]
                # Clean up partial files
                for f in [RESULTS_DIR / f"{test_id}.json",
                          RESULTS_DIR / f"{test_id}_progress.json"]:
                    if f.exists():
                        f.unlink()
                return True
        return False
    
    def reset_test(self, test_id):
        """Reset a test back to pending by clearing memory state and deleting files on disk."""
        removed = False
        with self.lock:
            if test_id in self.completed_tests:
                del self.completed_tests[test_id]
                removed = True
            if test_id in self.running_tests:
                # If running, we should cancel it first
                proc = self.running_tests[test_id].get("process")
                if proc:
                    try:
                        proc.terminate()
                    except Exception:
                        pass
                del self.running_tests[test_id]
                removed = True
        
        # Always delete the files on disk if they exist
        result_file = RESULTS_DIR / f"{test_id}.json"
        progress_file = RESULTS_DIR / f"{test_id}_progress.json"
        for f in [result_file, progress_file]:
            if f.exists():
                try:
                    f.unlink()
                    removed = True
                except Exception:
                    pass
        return removed or result_file.exists() or progress_file.exists()


# Initialize runner
runner = TestRunner()

# ── API Routes ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Serve the frontend."""
    resp = app.send_static_file("index.html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

@app.route("/api/tests", methods=["GET"])
def api_tests():
    """List all tests with their current status."""
    statuses = runner.get_all_statuses()
    suites = []
    for suite in TEST_SPEC["test_suites"]:
        suite_info = {
            "id": suite["id"],
            "name": suite["name"],
            "parallel": suite.get("parallel", True),
            "tests": [],
        }
        for test in suite["tests"]:
            tid = test["id"]
            test_info = {
                "id": tid,
                "name": test["name"],
                "type": test.get("type", "api_call"),
                "desc": test.get("desc", ""),
                "parallel": test.get("parallel", suite.get("parallel", True)),
                "official": test.get("official"),
                "tolerance": test.get("tolerance"),
                **statuses.get(tid, {"status": "pending", "progress": 0}),
            }
            suite_info["tests"].append(test_info)
        suites.append(suite_info)
    
    import copy
    safe_model = copy.deepcopy(CONFIG.get("model", {}))
    if "api_key" in safe_model:
        raw_key = safe_model["api_key"]
        if raw_key:
            safe_model["api_key"] = "••••••••"

    return jsonify({
        "model": {**safe_model, "api_url": safe_model.get("api_base", "")},
        "suites": suites,
        "non_parallel_locked": runner.is_non_parallel_running(),
    })

@app.route("/api/tests/<test_id>/run", methods=["POST"])
def api_run_test(test_id):
    """Start a test."""
    if test_id not in TESTS:
        return jsonify({"error": "Test not found"}), 404
    
    success, message = runner.start_test(test_id)
    if success:
        return jsonify({"status": "started", "test_id": test_id})
    else:
        return jsonify({"error": message}), 409

@app.route("/api/tests/<test_id>/status", methods=["GET"])
def api_test_status(test_id):
    """Get status of a specific test."""
    if test_id not in TESTS:
        return jsonify({"error": "Test not found"}), 404
    return jsonify(runner.get_test_status(test_id))

@app.route("/api/tests/<test_id>/cancel", methods=["POST"])
def api_cancel_test(test_id):
    """Cancel a running test."""
    if runner.cancel_test(test_id):
        return jsonify({"status": "cancelled"})
    return jsonify({"error": "Test not running"}), 404

@app.route("/api/tests/<test_id>/reset", methods=["POST"])
def api_reset_test(test_id):
    """Reset a completed test to pending."""
    runner.reset_test(test_id)
    return jsonify({"status": "reset"})

@app.route("/api/run-all", methods=["POST"])
def api_run_all():
    """Run all tests (clear old results first, then start all pending)."""
    started = []
    skipped = []
    
    # Clear all old results and progress files before re-running
    data = request.get_json(silent=True) or {}
    force = data.get("force", True)  # Default: clear old results
    if force:
        with runner.lock:
            runner.completed_tests.clear()
            for f in RESULTS_DIR.glob("*.json"):
                try:
                    f.unlink()
                except Exception:
                    pass
    
    # First pass: start all parallel tests
    for suite in TEST_SPEC["test_suites"]:
        for test in suite["tests"]:
            tid = test["id"]
            effective_parallel = test.get("parallel", suite.get("parallel", True))
            
            # Skip running/completed tests
            status = runner.get_test_status(tid)
            if status["status"] not in ("pending", "failed", "passed"):
                continue
            
            if effective_parallel:
                success, msg = runner.start_test(tid)
                if success:
                    started.append(tid)
                else:
                    skipped.append({"id": tid, "reason": msg})
            else:
                skipped.append({"id": tid, "reason": "Non-parallel test queued"})
    
    # Start one non-parallel test if possible
    if not runner.is_non_parallel_running():
        for suite in TEST_SPEC["test_suites"]:
            for test in suite["tests"]:
                tid = test["id"]
                effective_parallel = test.get("parallel", suite.get("parallel", True))
                if not effective_parallel:
                    status = runner.get_test_status(tid)
                    if status["status"] == "pending":
                        success, msg = runner.start_test(tid)
                        if success:
                            started.append(tid)
                            break
    
    return jsonify({
        "started": started,
        "skipped": skipped,
        "total": len(started) + len(skipped),
    })

@app.route("/api/clear-results", methods=["POST"])
def api_clear_results():
    """Clear all test results and remove files from disk."""
    with runner.lock:
        count = 0
        for f in RESULTS_DIR.glob("*.json"):
            try:
                f.unlink()
                count += 1
            except Exception:
                pass
        runner.completed_tests.clear()
    return jsonify({"status": "cleared", "files_removed": count})

@app.route("/api/results", methods=["GET"])
def api_results():
    """Get all test results."""
    results = {}
    with runner.lock:
        for tid, result in runner.completed_tests.items():
            results[tid] = result
    return jsonify({"results": results, "count": len(results)})

@app.route("/api/export", methods=["GET"])
def api_export():
    """Export results as xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500
    
    wb = openpyxl.Workbook()
    
    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "测试总览"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    pass_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    fail_fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
    pending_fill = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Title
    ws.merge_cells("A1:G1")
    model_name = CONFIG.get("model", {}).get("name", "LLM")
    ws["A1"] = f"{model_name} 基准测试报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Model info
    model = CONFIG.get("model", {})
    ws["A3"] = "模型名称"
    ws["B3"] = model.get("name", "N/A")
    ws["A4"] = "供应商"
    ws["B4"] = model.get("provider", "N/A")
    ws["A5"] = "最大上下文"
    ws["B5"] = f"{model.get('max_context', 'N/A')} tokens"
    
    # Headers
    headers = ["测试套件", "测试名称", "类型", "状态", "结果详情", "完成时间", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row = 8
    statuses = runner.get_all_statuses()
    for suite in TEST_SPEC["test_suites"]:
        for test in suite["tests"]:
            tid = test["id"]
            status_info = statuses.get(tid, {"status": "pending"})
            status = status_info.get("status", "pending")
            
            result_data = runner.completed_tests.get(tid, {})
            result = result_data.get("result", {})
            
            # Format result details: prefer details string, then metrics summary, then raw
            if isinstance(result, dict):
                detail_str = result.get("details", "")
                metrics = result.get("metrics", {})
                if isinstance(metrics, dict) and metrics:
                    # Summarize metrics: show pass/total for sub_tests, or ttft/otps for perf
                    if "sub_tests" in metrics:
                        st = metrics["sub_tests"]
                        detail_str += f" [{metrics.get('passed','?')}/{metrics.get('total','?')}]"
                        if isinstance(st, dict):
                            for sub_name, sub_val in st.items():
                                if isinstance(sub_val, dict):
                                    ok = "PASS" if sub_val.get("passed") else "FAIL"
                                    detail = sub_val.get("detail", "")
                                    desc = sub_val.get("desc", "")
                                    k25 = sub_val.get("k25_behavior", "")
                                    combined = "; ".join(filter(None, [detail, desc, k25]))
                                    detail_str += f"\n  [{ok}] {sub_name}: {combined[:150]}"
                    elif "ttft" in metrics:
                        t = metrics["ttft"]
                        detail_str += f" [P50={t.get('p50','?')}s P90={t.get('p90','?')}s]"
                    elif "success_rate" in metrics:
                        detail_str += f" [success_rate={metrics['success_rate']:.1f}%]"
                details = detail_str if detail_str else str(result)[:200]
            else:
                details = str(result) if result else ""
            
            completed_at = result_data.get("completed_at", "")
            if completed_at:
                try:
                    completed_at = datetime.fromisoformat(completed_at).strftime("%Y-%m-%d %H:%M")
                except ValueError:
                    pass
            
            status_map = {
                "passed": "通过",
                "failed": "失败",
                "pending": "待执行",
                "running": "执行中",
                "skipped": "已跳过",
            }
            
            ws.cell(row=row, column=1, value=suite["name"]).border = thin_border
            ws.cell(row=row, column=2, value=test["name"]).border = thin_border
            ws.cell(row=row, column=3, value=test.get("type", "")).border = thin_border
            status_cell = ws.cell(row=row, column=4, value=status_map.get(status, status))
            status_cell.border = thin_border
            if status == "passed":
                status_cell.fill = pass_fill
                status_cell.font = Font(color="FFFFFF")
            elif status == "failed":
                status_cell.fill = fail_fill
                status_cell.font = Font(color="FFFFFF")
            elif status == "pending":
                status_cell.fill = pending_fill
            
            # Remove illegal XML/Excel characters (e.g., tqdm \r or control characters)
            import re
            cleaned_details = re.sub(r'[\x00-\x08\x0b\x0c\x0e\x1f]', '', details)
            ws.cell(row=row, column=5, value=cleaned_details[:200]).border = thin_border
            ws.cell(row=row, column=6, value=completed_at).border = thin_border
            ws.cell(row=row, column=7, value=test.get("desc", "")).border = thin_border
            row += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 40
    
    # ── Benchmark Sheet ──
    ws2 = wb.create_sheet("精度测试")
    bench_headers = ["Benchmark", "官方分数", "实测分数", "差值", "容忍度", "结果"]
    for col, h in enumerate(bench_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    bench_row = 2
    for suite in TEST_SPEC["test_suites"]:
        if suite["id"] != "accuracy":
            continue
        for test in suite["tests"]:
            tid = test["id"]
            result_data = runner.completed_tests.get(tid, {})
            result = result_data.get("result", {})
            
            ws2.cell(row=bench_row, column=1, value=test["name"]).border = thin_border
            ws2.cell(row=bench_row, column=2, value=test.get("official", "N/A")).border = thin_border
            score = result.get("score")
            delta = result.get("delta")
            if score is None:
                metrics = result.get("metrics", {})
                if isinstance(metrics, dict) and "correct" in metrics and "total" in metrics:
                    total = metrics["total"]
                    correct = metrics["correct"]
                    if total > 0:
                        score = round((correct / total) * 100, 2)
                        official = test.get("official")
                        if isinstance(official, (int, float)):
                            delta = round(score - official, 2)
            
            if score is None:
                score = "N/A"
            if delta is None:
                delta = "N/A"

            ws2.cell(row=bench_row, column=3, value=score).border = thin_border
            ws2.cell(row=bench_row, column=4, value=delta).border = thin_border
            ws2.cell(row=bench_row, column=5, value=test.get("tolerance", 4.0)).border = thin_border
            status = result_data.get("status", "pending")
            ws2.cell(row=bench_row, column=6, value="通过" if status == "passed" else "失败" if status == "failed" else "待执行").border = thin_border
            bench_row += 1
    
    for col in "ABCDEF":
        ws2.column_dimensions[col].width = 20
    
    # ── Performance Sheet ──
    ws3 = wb.create_sheet("性能测试")
    perf_headers = ["测试项", "P50(s)", "P90(s)", "OTPS", "成功率", "结果"]
    for col, h in enumerate(perf_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    perf_row = 2
    for suite in TEST_SPEC["test_suites"]:
        if suite["id"] != "perf":
            continue
        for test in suite["tests"]:
            tid = test["id"]
            result_data = runner.completed_tests.get(tid, {})
            result = result_data.get("result", {})
            # Metrics are nested: result.metrics.ttft / result.metrics.otps / result.metrics.success_rate
            metrics = result.get("metrics", {})
            ttft_metrics = metrics.get("ttft", {})
            otps_metrics = metrics.get("otps", {})
            success_rate = metrics.get("success_rate", "N/A")
            
            ws3.cell(row=perf_row, column=1, value=test["name"]).border = thin_border
            ws3.cell(row=perf_row, column=2, value=f'{ttft_metrics.get("p50", "N/A")}' if ttft_metrics else "N/A").border = thin_border
            ws3.cell(row=perf_row, column=3, value=f'{ttft_metrics.get("p90", "N/A")}' if ttft_metrics else "N/A").border = thin_border
            ws3.cell(row=perf_row, column=4, value=f'{otps_metrics.get("avg", "N/A")}' if otps_metrics else "N/A").border = thin_border
            ws3.cell(row=perf_row, column=5, value=f'{success_rate}%' if isinstance(success_rate, (int,float)) else str(success_rate)).border = thin_border
            status = result_data.get("status", "pending")
            ws3.cell(row=perf_row, column=6, value="通过" if status == "passed" else "失败" if status == "failed" else "待执行").border = thin_border
            perf_row += 1
    
    for col in "ABCDEF":
        ws3.column_dimensions[col].width = 20
    
    # Save
    export_path = RESULTS_DIR / f"benchmark_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(export_path))
    
    model_name = CONFIG.get("model", {}).get("name", "LLM").lower().replace(" ", "_")
    return send_file(
        str(export_path),
        as_attachment=True,
        download_name=f"{model_name}_benchmark_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/api/stream/<test_id>")
def api_stream(test_id):
    """SSE endpoint for real-time test progress."""
    if test_id not in TESTS:
        return jsonify({"error": "Test not found"}), 404
    
    def generate():
        last_progress = -1
        last_message = None
        while True:
            is_running = False
            with runner.lock:
                is_running = test_id in runner.running_tests
            if is_running:
                runner._sync_progress_from_file(test_id)
                
            with runner.lock:
                if test_id in runner.running_tests:
                    rt = runner.running_tests[test_id]
                    progress = rt.get("progress", 0)
                    message = rt.get("message", "")
                    data = {
                        "status": "running",
                        "progress": progress,
                        "message": message,
                    }
                    if progress != last_progress or message != last_message:
                        last_progress = progress
                        last_message = message
                        yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
                elif test_id in runner.completed_tests:
                    result = runner.completed_tests[test_id]
                    data = {
                        "status": result.get("status", "completed"),
                        "progress": 100,
                        "result": result.get("result"),
                        "completed_at": result.get("completed_at"),
                    }
                    yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
                    break
                else:
                    data = {"status": "pending", "progress": 0}
                    yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
                    break
            time.sleep(1)
    
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

def verify_api_credentials(api_base, api_key):
    """Test the connectivity of the API using /models or /chat/completions."""
    # 1. Try /models GET
    url_models = api_base.rstrip('/') + '/models'
    try:
        req_models = urllib.request.Request(
            url_models,
            headers={
                "Authorization": f"Bearer {api_key}",
                "User-Agent": "Model-Benchmark-WebUI/1.0"
            }
        )
        with urllib.request.urlopen(req_models, timeout=5) as response:
            if response.getcode() == 200:
                return True, "连接成功"
    except urllib.error.HTTPError as e:
        if e.code in [401, 403]:
            return False, "API 密钥验证失败 (Unauthorized)"
    except urllib.error.URLError as e:
        return False, f"DNS 或网络连接错误: {str(e.reason)}"
    except Exception:
        pass

    # 2. Try /chat/completions POST
    url_chat = api_base.rstrip('/') + '/chat/completions'
    model_name = CONFIG.get("model", {}).get("name", "qwen3.7-max")
    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": "Ping"}],
        "max_tokens": 1
    }
    try:
        data = json.dumps(payload).encode('utf-8')
        req_chat = urllib.request.Request(
            url_chat,
            data=data,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "User-Agent": "Model-Benchmark-WebUI/1.0"
            },
            method="POST"
        )
        with urllib.request.urlopen(req_chat, timeout=8) as response:
            if response.getcode() == 200:
                return True, "连接成功"
            return False, f"接口返回 HTTP 状态码: {response.getcode()}"
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode('utf-8')
            err_json = json.loads(err_body)
            err_msg = err_json.get("error", {}).get("message", f"HTTP {e.code}")
        except Exception:
            err_msg = f"HTTP {e.code}"
        return False, f"接口错误: {err_msg}"
    except urllib.error.URLError as e:
        return False, f"DNS 或网络连接错误: {str(e.reason)}"
    except Exception as e:
        return False, f"测试请求发送失败: {str(e)}"


@app.route("/api/config", methods=["GET", "POST"])
def api_config():
    """Get or update configuration."""
    global CONFIG
    if request.method == "POST":
        try:
            data = request.get_json()
            if not data:
                return jsonify({"error": "Missing json body"}), 400
            
            api_base = data.get("api_base")
            api_key = data.get("api_key")
            
            if not api_base or not api_key:
                return jsonify({"error": "api_base and api_key are required"}), 400
            
            # If the submitted key matches the masked version of the current key, reuse the raw key!
            current_key = CONFIG.get("model", {}).get("api_key", "")
            if api_key == "••••••••":
                api_key = current_key
            
            # Test connectivity
            ok, msg = verify_api_credentials(api_base, api_key)
            if not ok:
                return jsonify({"error": f"绑定失败: {msg}"}), 400
            
            # Load current config.json
            if CONFIG_PATH.exists():
                with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
            else:
                cfg = {}
            
            if "model" not in cfg:
                cfg["model"] = {}
                
            cfg["model"]["api_base"] = api_base
            cfg["model"]["api_key"] = api_key
            
            # Write back to config.json
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)
                
            # Reload CONFIG in memory
            CONFIG = cfg
            
            return jsonify({"success": True, "message": "配置绑定成功"})
        except Exception as e:
            return jsonify({"error": f"更新失败: {str(e)}"}), 500
            
    import copy
    safe_config = copy.deepcopy(CONFIG)
    if "model" in safe_config and "api_key" in safe_config["model"]:
        raw_key = safe_config["model"]["api_key"]
        if raw_key:
            safe_config["model"]["api_key"] = "••••••••"
    return jsonify(safe_config)

# ── Main ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    print(f"🚀 Benchmark WebUI starting on http://0.0.0.0:{port}")
    print(f"📋 Loaded {len(TESTS)} tests from {len(TEST_SPEC['test_suites'])} suites")
    print(f"📊 Found {len(runner.completed_tests)} existing results")
    app.run(host="0.0.0.0", port=port, debug=debug, threaded=True)
